import os
from openpyxl import load_workbook
from flask import render_template
from flask import request
from flask import flash
from flask import redirect
from flask_login import login_required
from werkzeug.utils import secure_filename

from DataBase.DataBase import db
from DataBase.Sender import Senders, Documents
from DataBase.Transactions import MoneyTransactions, Banks, Reasons, Receivers, Origins
from config import ALLOWED_EXTENSIONS
from config import UPLOAD_FOLDER


@login_required
def xlsx_reader():
    profiles_list = Senders.query.all()
    bank_exists = db.session.query(db.exists().where(Banks.id == 1)).scalar()
    if bank_exists is False:
        bank_1 = Banks(bank_id=1, name="Nordea")
        bank_2 = Banks(bank_id=2, name="Danske Banken")
        db.session.add(bank_1)
        db.session.add(bank_2)
        db.session.commit()

    reasons_exists = db.session.query(db.exists().where(Reasons.id == 1)).scalar()
    if reasons_exists is False:
        reason_1 = Reasons(reason_id=1, reason_name="Stöd till familj/släkt")
        reason_2 = Reasons(reason_id=2, reason_name="Vård")
        reason_3 = Reasons(reason_id=3, reason_name="Studier")
        reason_4 = Reasons(reason_id=4, reason_name="Skickar till sig själv")
        reason_5 = Reasons(reason_id=5, reason_name="Investering")

        db.session.add(reason_1)
        db.session.add(reason_2)
        db.session.add(reason_3)
        db.session.add(reason_4)
        db.session.add(reason_5)
        db.session.commit()

    origin_of_money_exists = db.session.query(db.exists().where(Origins.id == 1)).scalar()
    if origin_of_money_exists is False:
        origin_of_money_1 = Origins(origin_id=1, origin_name="Lön")
        origin_of_money_2 = Origins(origin_id=2, origin_name="Besparingar")
        origin_of_money_3 = Origins(origin_id=3, origin_name="Lån")
        origin_of_money_4 = Origins(origin_id=4, origin_name="Försäljning av egendom")

        db.session.add(origin_of_money_1)
        db.session.add(origin_of_money_2)
        db.session.add(origin_of_money_3)
        db.session.add(origin_of_money_4)
        db.session.commit()

    if request.method == "POST":
        if 'transaktions_list' in request.form:
            tr_list = request.form
            tr_date = tr_list["date"]
            tr_bank_id = tr_list.getlist("bank")[0]
            tr_sender_name = tr_list.getlist("senders_name")
            tr_senders_ssn = tr_list.getlist("ssn")
            #tr_senders_document = tr_list.getlist("senders_document")
            tr_senders_telefon = tr_list.getlist("telefon")
            tr_amount = tr_list.getlist("amount")
            tr_reason = tr_list.getlist("reason")
            tr_origin = tr_list.getlist("origin")
            tr_senders_address = tr_list.getlist("address")
            tr_receiver_name = tr_list.getlist("receiver_name")
            tr_reference = tr_list.getlist("reference")
            tr_relationship = tr_list.getlist("relationship")
            tr_country = tr_list.getlist("country")
            expire_date_of_senders_document = tr_list.getlist("expire_date")
            type_of_senders_document = tr_list.getlist("type_of_senders_document")
            #file_of_senders_document = tr_list.getlist("file_of_senders_document")
            try:
                for i in range(len(tr_sender_name)):
                    receiver = Receivers(name=tr_receiver_name[i], relation_to_sender=tr_relationship[i],
                                         country=tr_country[i])
                    db.session.add(receiver)
                    db.session.flush()
                    sender_exists = db.session.query(db.exists().where(Senders.ssn == tr_senders_ssn[i])).scalar()
                    if sender_exists is True:
                        tr_sender = Senders.query.filter_by(ssn=tr_senders_ssn[i]).first()
                    else:
                        tr_sender_document = Documents(type_of_docs="",
                                                       expire_date="åååå-mm-dd",
                                                       file="")
                        db.session.add(tr_sender_document)
                        db.session.flush()
                        tr_sender = Senders(l_name="",
                                            f_name=tr_sender_name[i],
                                            ssn=tr_senders_ssn[i],
                                            address=tr_senders_address[i],
                                            phone_number=tr_senders_telefon[i],
                                            document_id=tr_sender_document.id,
                                            document=tr_sender_document)

                    tr_bank = Banks.query.filter_by(id=tr_bank_id).first()

                    reason = Reasons.query.filter_by(id=tr_reason[i]).first()
                    origin = Origins.query.filter_by(id=tr_origin[i]).first()

                    tr = MoneyTransactions(tr_amount=tr_amount[i],
                                           tr_reference=tr_reference[i],
                                           reason_id=tr_reason[i],
                                           reason=reason,
                                           tr_fee="1",
                                           tr_date=tr_date[0:10],
                                           bank_id=tr_bank_id,
                                           bank=tr_bank,
                                           sender=tr_sender,
                                           sender_ssn=tr_senders_ssn[i],
                                           receiver=receiver,
                                           receiver_id=receiver.id,
                                           origin_of_money=origin,
                                           origin_of_money_id=tr_origin[i])

                    db.session.add(tr_sender)
                    db.session.add(tr)
                    db.session.commit()
            except IndexError:
                pass
        else:
            pass

        if "xlsx_file" not in request.files:
            flash("No file part")
            return redirect(request.url)
        file = request.files["xlsx_file"]
        if file.filename == "":
            flash("No selected file")
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            data_from_file, file_date, total_amount = get_information_from_xlsx_file(
                UPLOAD_FOLDER + "/" + file.filename)
            # return redirect(url_for("xlsx_reader", data_from_file))
            reasons = Reasons.query.all()
            origins = Origins.query.all()
            return render_template("xlsx_reader.html",
                                   profiles=profiles_list,
                                   result=data_from_file,
                                   date=file_date,
                                   total_amount=total_amount,
                                   reasons=reasons,
                                   origins=origins)

    return render_template("xlsx_reader.html")


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def get_information_from_xlsx_file(file):
    wb = load_workbook(filename=file)
    ws = wb.active
    res = []
    datum = ws.cell(row=5, column=1).value
    total_amount = 0

    for i in range(1, ws.max_row + 1):
        if i > 18:
            # sender_name = ws.cell(row=i, column=1).value
            # receiver_name = ws.cell(row=i, column=3).value
            # reference = ws.cell(row=i, column=4).value
            total_amount += float(ws.cell(row=i, column=5).value)
            # avi = ws.cell(row=i, column=6).value
            # address = ws.cell(row=i, column=7).value
            res.append(
                {
                    "id": str(i - 18),
                    "sender_name": ws.cell(row=i, column=1).value,
                    "receiver_name": ws.cell(row=i, column=3).value,
                    "reference": ws.cell(row=i, column=4).value,
                    "amount": ws.cell(row=i, column=5).value,
                    "avi": ws.cell(row=i, column=6).value,
                    "address": ws.cell(row=i, column=7).value,
                }
            )
    return res, datum, total_amount


if __name__ == "__main__":
    pass
